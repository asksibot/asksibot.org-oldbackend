from flask_mail import Message
from datetime import datetime
import openpyxl

def handle_feedback_submission(app, feedback_data, recipient):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    feedback_message = "\n".join([f"{key}: {value}" for key, value in feedback_data.items()]) + f"\nTimestamp: {timestamp}"

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add the privacy statement
    worksheet['A1'] = 'Email (Optional)'
    worksheet['B1'] = 'Privacy Statement'
    worksheet['C1'] = 'Your privacy is important to us. All data collected is for internal review only and will not be shared.'

    # Add the feedback data to the worksheet
    row = 3
    for key, value in feedback_data.items():
        worksheet.cell(row=row, column=1, value=key)
        worksheet.cell(row=row, column=2, value=value)
        row += 1

    # Save the workbook to a file
    excel_file_path = 'feedback.xlsx'
    workbook.save(excel_file_path)

    msg = Message('Asksibot Feedback Submission', sender=app.config['MAIL_USERNAME'], recipients=[recipient])
    msg.body = feedback_message
    with open(excel_file_path, 'rb') as f:
        msg.attach(excel_file_path, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f.read())
    app.extensions['mail'].send(msg)

    # Clean up the temporary Excel file
    os.remove(excel_file_path)
